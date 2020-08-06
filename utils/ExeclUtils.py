# !/usr/bin/env python3
# -*- coding: utf-8 -*-

# ======================================================================
#   Copyright (C) 2020 liaozhimingandy@qq.com Ltd. All Rights Reserved.
#
#   @Author      : zhiming
#   @Project     : Demo
#   @File Name   : ExeclUtils.py	
#   @Created Date: 2020-06-15 17:46
#      @Software : PyCharm
#         @e-Mail: liaozhimingandy@qq.com
#   @Description :
#
# ======================================================================
import os

from tqdm import tqdm
import xlrd as xd
import openpyxl as xt
import pypinyin as ppy

from loguru import logger

class ExcelUtils(object):

    @staticmethod
    def excel_to_list(file_path: str) -> list:
        """
        从excel文件中读取数据并且转换成列表
        :param file_path: 文件路径
        :return: 列表
        """
        # 判断文件是否存在,不存在则直接跳出
        if not os.path.exists(file_path):
            print(r'不存在需要处理的文件<{0}>'.format(file_path))
            exit(1)
        sheet = xd.open_workbook(file_path).sheet_by_index(0)
        list_data = list()
        # 根据第一行数据作为字典的键
        excel_title = sheet.row_values(0, 0, sheet.ncols)
        # 2,数据提取到列表中
        for row in range(1, sheet.nrows):
            # 获取表格第一列和第二列里面值并且格式处理
            tmp_content = sheet.row_values(row, 0, sheet.ncols)
            # tmp_dict_data = dict(zip(excel_title, tmp_content))
            list_data.append(tmp_content)

        # list_data = list(map(deal_list_number, list_data))
        list_data.insert(0, excel_title)
        return list_data

    @staticmethod
    def execl_to_list_for_openpyxl(file_path_excel: str) -> dict:
        """
        通过excel文件名,将文件转成字典输出
        :param file_path_excel: excel文件路径名称
        :return: 字典
        """
        # 判断文件是否存在,不存在则直接跳出
        if not os.path.exists(file_path_excel):
            return {"error": 1, "msg": f"文件:{file_path_excel}不存在", "result": None}

        wb = xt.load_workbook(filename=file_path_excel)
        sheet = wb["Sheet1"]

        list_data = list()
        for row in sheet.rows:
            list_tmp = list()
            for cell in row:
                list_tmp.append(cell.value)

            list_data.append(list_tmp)

        return {"error": 0, "msg": f"已成功处理!", "data": list_data}

    @staticmethod
    def list_to_execl(list_data: list, file_path_to_save: str):
        """
        根据传入的列表参数,写入到指定的文件中,文件为excel
        :param list_data: 列表数据
        :param file_path_to_save: 文件路径
        :return:
        """
        # 判断文件是否存在,不存在则直接跳出
        if os.path.exists(file_path_to_save):
            print(r'已存在该文件<{0}>'.format(file_path_to_save))
            exit(1)

        wt = xt.Workbook()
        sheet = wt.create_sheet("result", 0)
        with tqdm(total=len(list_data)) as pbar:
            pbar.set_description("完成进度")
            for row in list_data:
                for n_col in range(len(row)):
                    sheet.cell(list_data.index(row) + 1, n_col + 1, row[n_col])
                pbar.update(1)

        wt.save(file_path_to_save)
        return {"error": 0, "msg": f"保存成功(路径名:{file_path_to_save})!", "data": None}


def deal_list_number(list_data: list, deal_col=1) -> list:
    list_data[deal_col] = int(list_data[deal_col])
    return list_data


def get_pinyin(str_content: str) -> str:
    return ''.join(ppy.lazy_pinyin(str_content, style=ppy.Style.FIRST_LETTER))\
        .upper().replace("／", "").replace("）", "").replace("（", "")


def main():
    file_path = r"D:\检验.xlsx"
    list_data = ExcelUtils.excel_to_list(file_path=file_path)

    with tqdm(total=len(list_data)) as pbar:
        pbar.set_description("list转换进度")
        for d in list_data:
            if list_data.index(d) == 0:
                continue
            d[0] = d[0].strip()
            d.append(get_pinyin(d[0]))
            d.append(list_data.index(d))
            # logger.info(d)
            pbar.update(1)

    # logger.info(list_data)
    file_path_to_save = "result.xlsx"
    ExcelUtils.list_to_execl(list_data, file_path_to_save)
    # print(ExcelUtils.execl_to_list_for_openpyxl(file_path_excel=file_path))


if __name__ == "__main__":
    main()
